"""
End-to-end smoke test for the HFFI pipeline.

Run with: pytest tests/ -v
or just:  python tests/test_pipeline.py
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import numpy as np
import pandas as pd

from hffi_core.scoring import HouseholdInputs, compute_household_hffi, hffi_score, risk_band
from hffi_core.weights import learn_weights_pca, learn_weights_logreg
from hffi_core.stress import apply_shock_scenarios, monte_carlo_stress
from hffi_core.recommendations import generate_recommendations
from hffi_core.validation import out_of_sample_eval, baseline_comparison, sensitivity_analysis
from data.synthetic import generate_households, add_distress_label, compute_components_for_population


def test_single_household():
    """A single household scoring should run end to end."""
    inputs = HouseholdInputs(
        monthly_income=6000, monthly_essential_expenses=3500, monthly_total_expenses=4500,
        liquid_savings=8000, total_debt=22000, monthly_debt_payment=800,
        portfolio_weights={"equity": 0.6, "bond": 0.3, "cash": 0.1},
        portfolio_volatility=0.16, expected_drawdown=0.20, rate_sensitivity=0.5,
    )
    macro = {"inflation_rate": 0.032, "fed_funds_rate": 0.05, "unemployment_rate": 0.038}
    result = compute_household_hffi(inputs, macro)
    assert 0 <= result.score <= 100
    assert result.band in {"Stable", "Moderate Fragility", "High Fragility", "Severe Fragility"}
    assert 0 <= result.distress_probability <= 1
    print(f"OK: Single household HFFI = {result.score:.1f} ({result.band})")


def test_score_monotonicity():
    """Higher debt → higher HFFI, all else equal."""
    base = HouseholdInputs(
        monthly_income=6000, monthly_essential_expenses=3500, monthly_total_expenses=4500,
        liquid_savings=8000, total_debt=10000, monthly_debt_payment=400,
        portfolio_weights={"equity": 0.6, "bond": 0.4},
        portfolio_volatility=0.15, expected_drawdown=0.20, rate_sensitivity=0.5,
    )
    macro = {"inflation_rate": 0.032, "fed_funds_rate": 0.05, "unemployment_rate": 0.038}
    score_low = compute_household_hffi(base, macro).score

    high_debt = HouseholdInputs(**{**base.__dict__, "total_debt": 80000, "monthly_debt_payment": 1500})
    score_high = compute_household_hffi(high_debt, macro).score
    assert score_high > score_low, "Higher debt must increase HFFI"
    print(f"OK: Monotonicity in debt — low={score_low:.1f}, high={score_high:.1f}")


def test_population_pipeline():
    """Generate a population, compute components, learn weights, validate."""
    print("\n=== Population pipeline ===")
    df = generate_households(n=2000, seed=42)
    macro = {"inflation_rate": 0.032, "fed_funds_rate": 0.05, "unemployment_rate": 0.038}
    components = compute_components_for_population(df, macro)
    y = add_distress_label(df, components, base_rate=0.15)
    print(f"Population: {len(df)} households, {y.sum()} in distress ({y.mean():.1%})")

    # Weight learning — PCA
    pca_weights, pca_diag = learn_weights_pca(components)
    print("\nPCA-learned weights:")
    for k, v in pca_weights.items():
        print(f"  {k}: {v:.3f}")
    print(f"  PC1 explained variance: {pca_diag['explained_variance_ratio'][0]:.3f}")

    # Weight learning — Logistic regression
    lr_weights, lr_diag = learn_weights_logreg(components, y)
    print("\nLogReg-learned weights:")
    for k, v in lr_weights.items():
        print(f"  {k}: {v:.3f}")
    print(f"  CV5 AUC: {lr_diag['cv5_auc']:.3f}")

    # Validation
    print("\n=== Validation ===")
    oos = out_of_sample_eval(components, y, weights=lr_weights)
    print(f"Out-of-sample AUC: {oos['mean_auc']:.3f} ± {oos['std_auc']:.3f}")
    print(f"Mean Brier score:  {oos['mean_brier']:.3f}")

    print("\nBaseline comparison:")
    bc = baseline_comparison(components, y, weights=lr_weights)
    print(bc.to_string(index=False))

    print("\nSensitivity analysis (±20% on each weight):")
    sa = sensitivity_analysis(components, weights=lr_weights, perturbation=0.20)
    print(sa.to_string(index=False))


def test_stress_simulation():
    inputs = HouseholdInputs(
        monthly_income=6000, monthly_essential_expenses=3500, monthly_total_expenses=4500,
        liquid_savings=8000, total_debt=22000, monthly_debt_payment=800,
        portfolio_weights={"equity": 0.6, "bond": 0.3, "cash": 0.1},
        portfolio_volatility=0.16, expected_drawdown=0.20, rate_sensitivity=0.5,
    )
    macro = {"inflation_rate": 0.032, "fed_funds_rate": 0.05, "unemployment_rate": 0.038}
    print("\n=== Stress scenarios ===")
    df = apply_shock_scenarios(inputs, macro)
    print(df.to_string())

    print("\n=== Monte Carlo ===")
    mc = monte_carlo_stress(inputs, macro, n_sims=1000)
    print(f"Mean post-shock HFFI: {mc.mean:.1f}")
    print(f"5th–95th percentile:  {mc.p05:.1f} – {mc.p95:.1f}")
    print(f"P(severe fragility):  {mc.prob_severe:.1%}")


def test_recommendations():
    inputs = HouseholdInputs(
        monthly_income=4000, monthly_essential_expenses=3500, monthly_total_expenses=3800,
        liquid_savings=500, total_debt=45000, monthly_debt_payment=1200,
        portfolio_weights={"equity": 0.95, "bond": 0.05},
        portfolio_volatility=0.25, expected_drawdown=0.40, rate_sensitivity=0.9,
    )
    macro = {"inflation_rate": 0.05, "fed_funds_rate": 0.06, "unemployment_rate": 0.05}
    result = compute_household_hffi(inputs, macro)
    print(f"\n=== Recommendations for fragile household (HFFI={result.score:.1f}, {result.band}) ===")
    recs = generate_recommendations(result, inputs)
    for r in recs["actions"]:
        print(f"  [P{r.priority}] {r.action}")
        print(f"        ({r.component}) — {r.expected_impact}")
    print(f"\n  Allocation rationale: {recs['rationale']}")
    print(f"  Allocation: {recs['allocation']}")


def test_investment_plan():
    """Investment plan generates sensible Monte Carlo wealth projections."""
    from hffi_core.investment_plan import build_investment_plan, compare_portfolios, PORTFOLIO_PARAMS
    plan = build_investment_plan("Balanced", 10, 10000, 500, 0.03, n_sims=1000)
    assert plan.horizon_years == 10
    assert plan.portfolio == "Balanced"
    assert plan.final_p5 < plan.final_p50 < plan.final_p95, "Percentile band must be ordered"
    assert plan.final_p50 > 10000, "Median wealth should exceed initial capital"
    assert len(plan.yearly_schedule) == 11, "Should have schedule for years 0..10"
    print(f"OK: Investment plan year-10 median = ${plan.final_p50:,.0f} "
          f"(p5-p95: ${plan.final_p5:,.0f}-${plan.final_p95:,.0f})")

    # Comparison must include all four portfolios
    comp = compare_portfolios(10, 10000, 500)
    assert len(comp) == len(PORTFOLIO_PARAMS), "Comparison must cover all portfolios"
    # Aggressive Growth should have widest dispersion
    agg = comp[comp["portfolio"] == "Aggressive Growth"].iloc[0]
    cons = comp[comp["portfolio"] == "Conservative"].iloc[0]
    agg_spread = agg["95th_pctile_value_y10"] - agg["5th_pctile_value_y10"]
    cons_spread = cons["95th_pctile_value_y10"] - cons["5th_pctile_value_y10"]
    assert agg_spread > cons_spread, "Aggressive must have wider 5th-95th band than Conservative"
    print(f"OK: Aggressive band ${agg_spread:,.0f} > Conservative band ${cons_spread:,.0f}")


def test_portfolio_recommender():
    """Portfolio recommender respects HFFI-conditional gating."""
    from hffi_core.market_recommender import (
        score_portfolios, allowed_portfolios, fragility_target_profile,
    )
    # Stable household should be offered Aggressive Growth
    assert "Aggressive Growth" in allowed_portfolios(15)
    # Severe fragility should ONLY allow Conservative
    assert allowed_portfolios(90) == ["Conservative"]
    # Targets should monotonically tighten as HFFI rises
    t_stable = fragility_target_profile(15)
    t_severe = fragility_target_profile(90)
    assert t_severe["preferred_volatility"] < t_stable["preferred_volatility"]
    assert t_severe["preferred_drawdown"] < t_stable["preferred_drawdown"]
    assert t_severe["preferred_liquidity"] > t_stable["preferred_liquidity"]
    # Score ordering at moderate HFFI should be deterministic
    scored = score_portfolios(55)
    assert "suitability_score" in scored.columns
    assert len(scored) == 3, "HFFI=55 should allow exactly 3 portfolios"
    print(f"OK: Portfolio recommender — top choice at HFFI=55 is {scored.iloc[0]['portfolio']}")


def test_market_suitability():
    """Market suitability scorer ranks low-vol assets higher for fragile households."""
    from hffi_core.market_recommender import score_markets_for_household
    snapshot = pd.DataFrame([
        {"market": "QQQ", "market_return": 0.018, "market_volatility": 0.22,
         "market_drawdown": -0.28, "market_sharpe_proxy": 0.082,
         "momentum_score": 0.020, "safety_score": -0.07},
        {"market": "AGG", "market_return": 0.003, "market_volatility": 0.05,
         "market_drawdown": -0.06, "market_sharpe_proxy": 0.06,
         "momentum_score": 0.003, "safety_score": -0.02},
    ])
    # Fragile household, risk-off → bonds should score better than QQQ
    s_fragile = score_markets_for_household(
        hffi=85, risk_band="Severe Fragility",
        debt_service_ratio=0.40, macro_stress_index=0.10, liquidity_buffer_6m=0.05,
        market_snapshot=snapshot, risk_off_regime=True,
    )
    agg_score = s_fragile[s_fragile["market"] == "AGG"]["suitability_score"].iloc[0]
    qqq_score = s_fragile[s_fragile["market"] == "QQQ"]["suitability_score"].iloc[0]
    assert agg_score > qqq_score, (
        f"Bonds should beat QQQ for fragile/risk-off household: "
        f"AGG={agg_score:.3f} QQQ={qqq_score:.3f}")
    print(f"OK: Fragile/risk-off AGG ({agg_score:+.2f}) > QQQ ({qqq_score:+.2f})")


def test_holdings_allocation_advisor():
    """Detailed Equity/Bond/Cash holdings produce allocation and action signals."""
    from hffi_core.portfolio_advisor import (
        PortfolioHolding, allocation_weights_from_holdings, build_holdings_dataframe,
        recommend_holding_actions, target_core_allocation,
    )

    holdings = [
        PortfolioHolding("equity", "AAPL", 10, 150, "Apple"),
        PortfolioHolding("bond", "AGG", 20, 95, "US Aggregate Bonds"),
        PortfolioHolding("cash", "CASH", 1000, 1, "Cash"),
    ]
    prices = {"AAPL": 180, "AGG": 100, "CASH": 1}
    holdings_df = build_holdings_dataframe(holdings, prices)
    weights = allocation_weights_from_holdings(holdings_df)
    assert round(sum(weights.values()), 6) == 1.0
    assert weights["equity"] > 0
    assert weights["bond"] > 0
    assert weights["cash"] > 0

    target = target_core_allocation({
        "cash_emergency": 0.20,
        "short_bonds": 0.30,
        "us_equities": 0.40,
        "intl_equities": 0.10,
    })
    actions = recommend_holding_actions(holdings_df, target, hffi=55)
    assert actions
    assert {a.ticker for a in actions} == {"AAPL", "AGG", "CASH"}
    print("OK: Holdings advisor calculates allocation and holding-level actions")


def test_data_science_recommender():
    """ML suitability layer trains, scores markets, and returns explanations."""
    from hffi_core.ds_recommender import (
        engineer_household_features, score_market_recommendations,
        train_suitability_model,
    )
    from hffi_core.portfolio_advisor import target_core_allocation

    inputs = HouseholdInputs(
        monthly_income=7000, monthly_essential_expenses=3200, monthly_total_expenses=4500,
        liquid_savings=15000, total_debt=20000, monthly_debt_payment=600,
        portfolio_weights={"equity": 0.10, "bond": 0.39, "cash": 0.51},
        portfolio_volatility=0.13, expected_drawdown=0.18, rate_sensitivity=0.4,
    )
    macro = {"inflation_rate": 0.032, "fed_funds_rate": 0.05, "unemployment_rate": 0.038}
    result = compute_household_hffi(inputs, macro)
    model = train_suitability_model()
    household_features = engineer_household_features(inputs, result, macro)
    market_scores = pd.DataFrame([
        {"market": "MSFT", "ticker": "MSFT", "name": "Microsoft", "category": "equity",
         "market_return": 0.02, "market_volatility": 0.07, "market_drawdown": -0.08,
         "market_sharpe_proxy": 0.28, "momentum_score": 0.02, "safety_score": -0.03,
         "suitability_score": 0.35},
        {"market": "AGG", "ticker": "AGG", "name": "US Aggregate Bonds", "category": "bond",
         "market_return": 0.003, "market_volatility": 0.03, "market_drawdown": -0.04,
         "market_sharpe_proxy": 0.10, "momentum_score": 0.003, "safety_score": -0.02,
         "suitability_score": 0.25},
    ])
    target = target_core_allocation({
        "cash_emergency": 0.20,
        "short_bonds": 0.30,
        "us_equities": 0.30,
        "intl_equities": 0.10,
    })
    scored = score_market_recommendations(
        model, household_features, market_scores, target,
        {"equity": 0.10, "bond": 0.39, "cash": 0.51},
    )
    assert not scored.empty
    assert {"ml_probability", "ds_score", "comment"}.issubset(scored.columns)
    assert scored["ds_score"].between(0, 1).all()
    print("OK: Data-science recommender trains, scores, and explains recommendations")


def test_evidence_engine():
    """Evidence Lab produces counterfactuals, confidence, and model-card rows."""
    from hffi_core.ds_recommender import (
        engineer_household_features, score_market_recommendations,
        train_suitability_model,
    )
    from hffi_core.evidence_engine import (
        build_counterfactual_table, build_decision_evidence_table,
        build_feature_evidence_table, build_model_card_table,
    )
    from hffi_core.portfolio_advisor import target_core_allocation

    inputs = HouseholdInputs(
        monthly_income=6500, monthly_essential_expenses=3300, monthly_total_expenses=4600,
        liquid_savings=5000, total_debt=30000, monthly_debt_payment=900,
        portfolio_weights={"equity": 0.65, "bond": 0.25, "cash": 0.10},
        portfolio_volatility=0.18, expected_drawdown=0.24, rate_sensitivity=0.6,
    )
    macro = {"inflation_rate": 0.035, "fed_funds_rate": 0.05, "unemployment_rate": 0.04}
    result = compute_household_hffi(inputs, macro)
    counter = build_counterfactual_table(inputs, macro, result)
    assert not counter.empty
    assert {"Action", "HFFI improvement", "Why it matters"}.issubset(counter.columns)

    model = train_suitability_model()
    features = engineer_household_features(inputs, result, macro)
    target = target_core_allocation({
        "cash_emergency": 0.20, "short_bonds": 0.30,
        "us_equities": 0.30, "intl_equities": 0.10,
    })
    market_scores = pd.DataFrame([{
        "market": "SPY", "ticker": "SPY", "name": "S&P 500", "category": "index",
        "market_return": 0.01, "market_volatility": 0.06, "market_drawdown": -0.08,
        "market_sharpe_proxy": 0.16, "momentum_score": 0.01, "safety_score": -0.03,
        "suitability_score": 0.20,
    }])
    ds_signals = score_market_recommendations(
        model, features, market_scores, target,
        {"equity": 0.65, "bond": 0.25, "cash": 0.10},
    )
    evidence = build_decision_evidence_table(ds_signals=ds_signals, hffi=result.score)
    assert not evidence.empty
    assert evidence["Confidence"].between(0, 1).all()
    feature_evidence = build_feature_evidence_table(model, features)
    assert not feature_evidence.empty
    model_card = build_model_card_table(model)
    assert {"Purpose", "Security"}.issubset(set(model_card["Section"]))
    print("OK: Evidence Lab builds counterfactuals, confidence, feature evidence, and model card")


def test_excel_report_generation():
    """Full Excel report end-to-end."""
    from hffi_core.investment_plan import build_investment_plan, compare_portfolios
    from hffi_core.market_recommender import score_markets_for_household, generate_trade_signals
    from hffi_core.report_generator import generate_report
    from openpyxl import load_workbook

    inputs = HouseholdInputs(
        monthly_income=6000, monthly_essential_expenses=3500, monthly_total_expenses=4500,
        liquid_savings=8000, total_debt=22000, monthly_debt_payment=800,
        portfolio_weights={"equity": 0.6, "bond": 0.3, "cash": 0.1},
        portfolio_volatility=0.16, expected_drawdown=0.20, rate_sensitivity=0.5,
    )
    macro = {"inflation_rate": 0.032, "fed_funds_rate": 0.05, "unemployment_rate": 0.038,
             "vix": 14.5, "treasury_10y": 0.044, "treasury_2y": 0.048,
             "yield_curve_spread": -0.004, "mortgage_rate": 0.072, "timestamp": "test"}
    result = compute_household_hffi(inputs, macro)
    plan = build_investment_plan("Balanced", 10, 10000, 500, 0.03, n_sims=500)
    out = Path("reports/test_pipeline_report.xlsx")
    generate_report(
        output_path=out, fragility_result=result, investment_plan=plan, macro=macro,
        portfolio_choice="Balanced", initial_capital=10000,
        monthly_contribution=500, horizon_years=10,
        portfolio_comparison=compare_portfolios(10, 10000, 500),
        stress_scenarios=apply_shock_scenarios(inputs, macro),
        recommendations=generate_recommendations(result, inputs),
    )
    assert out.exists(), "Report file must exist"
    wb = load_workbook(out)
    expected_sheets = {"Executive Summary", "Component Breakdown", "Investment Plan",
                       "Portfolio Comparison", "Stress Scenarios", "Recommendations",
                       "Macro Snapshot"}
    assert expected_sheets.issubset(wb.sheetnames), \
        f"Missing sheets: {expected_sheets - set(wb.sheetnames)}"
    print(f"OK: Excel report generated ({out.stat().st_size:,} bytes, {len(wb.sheetnames)} sheets)")


def test_chatbot_fallback():
    """Chatbot rule-based fallback returns sensible answers without API key."""
    from hffi_core.chatbot import _rule_based_response
    inputs = HouseholdInputs(
        monthly_income=6000, monthly_essential_expenses=3500, monthly_total_expenses=4500,
        liquid_savings=8000, total_debt=22000, monthly_debt_payment=800,
        portfolio_weights={"equity": 0.6, "bond": 0.3, "cash": 0.1},
        portfolio_volatility=0.16, expected_drawdown=0.20, rate_sensitivity=0.5,
    )
    macro = {"inflation_rate": 0.032, "fed_funds_rate": 0.05, "unemployment_rate": 0.038}
    result = compute_household_hffi(inputs, macro)
    state = {"fragility_result": result, "inputs": inputs, "macro": macro,
             "portfolio_choice": "Balanced"}
    for q in ["What is HFFI?", "Tell me about my liquidity",
              "Why is my debt so high?", "Should I buy NVDA?"]:
        reply = _rule_based_response(q, state)
        assert len(reply) > 50, f"Reply too short: {reply}"
    print("OK: Chatbot rule-based fallback responds to common queries")


if __name__ == "__main__":
    test_single_household()
    test_score_monotonicity()
    test_stress_simulation()
    test_recommendations()
    test_population_pipeline()
    test_investment_plan()
    test_portfolio_recommender()
    test_market_suitability()
    test_holdings_allocation_advisor()
    test_data_science_recommender()
    test_evidence_engine()
    test_excel_report_generation()
    test_chatbot_fallback()
    print("\nOK: All tests passed.")
