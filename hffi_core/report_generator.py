"""
Excel report generator.

Produces a multi-sheet Excel workbook containing:
    - Sheet 1: Executive Summary (HFFI score, band, key metrics)
    - Sheet 2: Component Breakdown (L, D, E, P, M with explanations)
    - Sheet 3: Investment Plan (year-by-year wealth schedule)
    - Sheet 4: Portfolio Comparison (4 portfolios side by side)
    - Sheet 5: Trade Signals (specific assets to buy/hold/avoid)
    - Sheet 6: Stress Scenarios (HFFI under 7 named shocks)
    - Sheet 7: Recommendations (rule-based actions)
    - Sheet 8: Macro Snapshot (current FRED indicators)

Color coding follows financial-modeling conventions:
    blue  = hardcoded inputs
    black = formulas / derived
    green = cross-sheet links
"""

from __future__ import annotations
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference

logger = logging.getLogger(__name__)


# Style palette
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_LABEL = Font(name="Calibri", size=11, bold=True)
FONT_INPUT = Font(name="Calibri", size=11, color="0000FF")
FONT_BODY = Font(name="Calibri", size=11)

FILL_TITLE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_HEADER = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
FILL_BAND_STABLE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_BAND_MODERATE = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_BAND_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_BAND_SEVERE = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _fill_for_band(band: str) -> PatternFill:
    return {
        "Stable":              FILL_BAND_STABLE,
        "Moderate Fragility":  FILL_BAND_MODERATE,
        "High Fragility":      FILL_BAND_HIGH,
        "Severe Fragility":    FILL_BAND_SEVERE,
    }.get(band, FILL_BAND_MODERATE)


def _set_title(ws, row: int, text: str, span: int = 6):
    """Write a section title row with the project's banner styling."""
    ws.cell(row=row, column=1, value=text).font = FONT_TITLE
    ws.cell(row=row, column=1).fill = FILL_TITLE
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _autosize_columns(ws, max_width: int = 50):
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if hasattr(cell, "column_letter"):
                col_letter = cell.column_letter
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


# --------------------------------------------------------------------------- #
# Sheet builders
# --------------------------------------------------------------------------- #
def _sheet_summary(wb: Workbook, ctx: dict):
    ws = wb.create_sheet("Executive Summary")
    _set_title(ws, 1, "HFFI TERMINAL — INVESTMENT REPORT", span=4)
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = FONT_BODY

    # Headline metrics
    ws.cell(row=4, column=1, value="HEADLINE FRAGILITY METRICS").font = FONT_HEADER
    ws.cell(row=4, column=1).fill = FILL_HEADER

    rows = [
        ("HFFI Score",            f"{ctx['hffi']:.1f} / 100",   "0–100 composite fragility index"),
        ("Risk Band",             ctx["band"],                   "Stable / Moderate / High / Severe"),
        ("12-month Distress Probability", f"{ctx['distress_prob']:.1%}", "Calibrated logistic"),
        ("Recommended Portfolio", ctx["portfolio_choice"],       "HFFI-conditional best fit"),
        ("Investment Horizon",    f"{ctx['horizon_years']} years", "User-selected"),
        ("Initial Capital",       f"${ctx['initial_capital']:,.0f}", ""),
        ("Monthly Contribution",  f"${ctx['monthly_contribution']:,.0f}", ""),
    ]
    for i, (label, value, note) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=label).font = FONT_LABEL
        ws.cell(row=i, column=2, value=value).font = FONT_BODY
        ws.cell(row=i, column=3, value=note).font = Font(name="Calibri", size=10, italic=True, color="666666")

    # Highlight band cell with color
    band_cell = ws.cell(row=6, column=2)
    band_cell.fill = _fill_for_band(ctx["band"])
    band_cell.font = FONT_LABEL

    # Investment plan headline
    ws.cell(row=14, column=1, value="INVESTMENT OUTCOME PROJECTION").font = FONT_HEADER
    ws.cell(row=14, column=1).fill = FILL_HEADER
    plan = ctx["investment_plan"]
    plan_rows = [
        ("Expected return (annualized)", f"{plan.expected_return:.1%}"),
        ("Volatility (annualized)",      f"{plan.volatility:.1%}"),
        (f"Median wealth at year {plan.horizon_years}", f"${plan.final_p50:,.0f}"),
        (f"5th percentile (downside)",   f"${plan.final_p5:,.0f}"),
        (f"95th percentile (upside)",    f"${plan.final_p95:,.0f}"),
        (f"Expected (mean)",             f"${plan.final_expected:,.0f}"),
    ]
    for i, (label, value) in enumerate(plan_rows, start=15):
        ws.cell(row=i, column=1, value=label).font = FONT_LABEL
        ws.cell(row=i, column=2, value=value).font = FONT_BODY

    ws.cell(row=23, column=1, value="SUMMARY").font = FONT_HEADER
    ws.cell(row=23, column=1).fill = FILL_HEADER
    ws.cell(row=24, column=1, value=plan.summary).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=24, start_column=1, end_row=27, end_column=4)
    ws.row_dimensions[24].height = 20
    _autosize_columns(ws, max_width=80)


def _sheet_components(wb: Workbook, ctx: dict):
    ws = wb.create_sheet("Component Breakdown")
    _set_title(ws, 1, "FRAGILITY COMPONENT BREAKDOWN", span=4)

    headers = ["Component", "Value (0-1)", "Points contributed", "Interpretation"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font, c.fill = FONT_HEADER, FILL_HEADER

    result = ctx["fragility_result"]
    interpretations = {
        "L": ("Liquidity",   "How many months of essentials savings cover. Higher = worse."),
        "D": ("Debt",        "Combined debt-service ratio + leverage. Higher = worse."),
        "E": ("Expenses",    "Share of essential spending. Higher = less buffer to cut."),
        "P": ("Portfolio",   "Volatility + concentration + drawdown exposure."),
        "M": ("Macro",       "Current macro regime stress weighted by your sensitivity."),
    }
    rows_data = [
        ("Liquidity (L)",  result.L, result.contributions["Liquidity"],  interpretations["L"][1]),
        ("Debt (D)",       result.D, result.contributions["Debt"],       interpretations["D"][1]),
        ("Expenses (E)",   result.E, result.contributions["Expenses"],   interpretations["E"][1]),
        ("Portfolio (P)",  result.P, result.contributions["Portfolio"],  interpretations["P"][1]),
        ("Macro (M)",      result.M, result.contributions["Macro"],      interpretations["M"][1]),
        ("Interaction (L×D)", result.interaction_LD, result.contributions["Interaction (L×D)"],
         "Joint stress of low liquidity AND high debt — danger zone."),
    ]
    for i, (name, val, pts, note) in enumerate(rows_data, start=4):
        ws.cell(row=i, column=1, value=name).font = FONT_LABEL
        ws.cell(row=i, column=2, value=round(val, 3)).font = FONT_BODY
        ws.cell(row=i, column=3, value=round(pts, 2)).font = FONT_BODY
        ws.cell(row=i, column=4, value=note).font = Font(name="Calibri", size=10, italic=True)

    # Total row
    total_row = 4 + len(rows_data)
    ws.cell(row=total_row, column=1, value="TOTAL HFFI SCORE").font = FONT_LABEL
    ws.cell(row=total_row, column=3, value=round(result.score, 1)).font = FONT_LABEL
    ws.cell(row=total_row, column=3).fill = FILL_HEADER
    ws.cell(row=total_row, column=3).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

    # Bar chart of components
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Fragility component contributions to HFFI"
    chart.y_axis.title = "Component"
    chart.x_axis.title = "Points contributed"
    data = Reference(ws, min_col=3, min_row=3, max_row=4 + len(rows_data) - 1)
    cats = Reference(ws, min_col=1, min_row=4, max_row=4 + len(rows_data) - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, "F3")
    _autosize_columns(ws, max_width=60)


def _sheet_investment_plan(wb: Workbook, ctx: dict):
    ws = wb.create_sheet("Investment Plan")
    _set_title(ws, 1, f"INVESTMENT PLAN — {ctx['horizon_years']} YEAR HORIZON", span=7)

    plan = ctx["investment_plan"]
    sched = plan.yearly_schedule

    # Inputs panel
    ws.cell(row=3, column=1, value="ASSUMPTIONS").font = FONT_HEADER
    ws.cell(row=3, column=1).fill = FILL_HEADER
    inputs = [
        ("Portfolio template",  plan.portfolio),
        ("Expected return",     plan.expected_return),
        ("Volatility",          plan.volatility),
        ("Initial capital",     plan.initial_capital),
        ("Monthly contribution",plan.monthly_contribution),
        ("Annual contribution growth", plan.annual_contribution_growth),
        ("Monte Carlo simulations", plan.n_simulations),
    ]
    for i, (label, val) in enumerate(inputs, start=4):
        ws.cell(row=i, column=1, value=label).font = FONT_LABEL
        c = ws.cell(row=i, column=2, value=val)
        c.font = FONT_INPUT
        if isinstance(val, float):
            if "return" in label.lower() or "volatility" in label.lower() or "growth" in label.lower():
                c.number_format = "0.0%"
            else:
                c.number_format = "$#,##0"

    # Schedule table
    table_start = 13
    ws.cell(row=table_start, column=1, value="YEAR-BY-YEAR PROJECTION").font = FONT_HEADER
    ws.cell(row=table_start, column=1).fill = FILL_HEADER

    cols = ["Year", "Cumulative contributions", "5th percentile", "Median (50th)",
            "Expected value", "95th percentile", "Growth above contributions"]
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=table_start + 1, column=j, value=c)
        cell.font, cell.fill = FONT_HEADER, FILL_HEADER

    for i, (_, row) in enumerate(sched.iterrows(), start=table_start + 2):
        ws.cell(row=i, column=1, value=int(row["year"]))
        ws.cell(row=i, column=2, value=row["cumulative_contribution"]).number_format = "$#,##0"
        ws.cell(row=i, column=3, value=row["p5"]).number_format = "$#,##0"
        ws.cell(row=i, column=4, value=row["p50"]).number_format = "$#,##0"
        ws.cell(row=i, column=5, value=row["expected_value"]).number_format = "$#,##0"
        ws.cell(row=i, column=6, value=row["p95"]).number_format = "$#,##0"
        ws.cell(row=i, column=7, value=row["growth_above_contributions"]).number_format = "$#,##0;($#,##0)"

    # Line chart
    chart = LineChart()
    chart.title = f"{plan.portfolio} portfolio — wealth projection"
    chart.x_axis.title = "Year"
    chart.y_axis.title = "Portfolio value ($)"
    end_row = table_start + 1 + len(sched)
    data = Reference(ws, min_col=3, min_row=table_start + 1, max_row=end_row, max_col=6)
    cats = Reference(ws, min_col=1, min_row=table_start + 2, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 18
    ws.add_chart(chart, "I3")
    _autosize_columns(ws, max_width=30)


def _sheet_portfolio_comparison(wb: Workbook, ctx: dict):
    if ctx.get("portfolio_comparison") is None or ctx["portfolio_comparison"].empty:
        return
    ws = wb.create_sheet("Portfolio Comparison")
    _set_title(ws, 1, "ALL PORTFOLIOS — SIDE BY SIDE", span=7)
    df = ctx["portfolio_comparison"]
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=3, column=j, value=col)
        c.font, c.fill = FONT_HEADER, FILL_HEADER
    for i, (_, row) in enumerate(df.iterrows(), start=4):
        for j, col in enumerate(df.columns, start=1):
            c = ws.cell(row=i, column=j, value=row[col])
            if isinstance(row[col], (int, float)):
                if "return" in col.lower() or "volatility" in col.lower():
                    c.number_format = "0.0%"
                else:
                    c.number_format = "$#,##0"
    _autosize_columns(ws, max_width=30)


def _sheet_trade_signals(wb: Workbook, ctx: dict):
    if not ctx.get("trade_signals"):
        return
    ws = wb.create_sheet("Trade Signals")
    _set_title(ws, 1, "RECOMMENDED ASSETS — BUY / HOLD / AVOID", span=7)

    headers = ["Action", "Ticker", "Name", "Target Allocation %", "Suggested Horizon",
               "Suitability", "Confidence", "Rationale"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font, c.fill = FONT_HEADER, FILL_HEADER

    for i, sig in enumerate(ctx["trade_signals"], start=4):
        ws.cell(row=i, column=1, value=sig.action)
        ws.cell(row=i, column=2, value=sig.ticker)
        ws.cell(row=i, column=3, value=sig.name)
        ws.cell(row=i, column=4, value=round(sig.target_allocation_pct, 2)).number_format = "0.00\"%\""
        ws.cell(row=i, column=5, value=sig.suggested_horizon)
        ws.cell(row=i, column=6, value=round(sig.suitability, 3))
        ws.cell(row=i, column=7, value=round(sig.confidence, 3))
        ws.cell(row=i, column=8, value=sig.rationale)
        ws.cell(row=i, column=8).alignment = Alignment(wrap_text=True)
        # Color the action cell
        action_cell = ws.cell(row=i, column=1)
        if sig.action == "BUY":
            action_cell.fill = FILL_BAND_STABLE
        elif sig.action == "HOLD":
            action_cell.fill = FILL_BAND_MODERATE
        elif sig.action == "AVOID":
            action_cell.fill = FILL_BAND_HIGH
    _autosize_columns(ws, max_width=70)


def _sheet_stress(wb: Workbook, ctx: dict):
    if ctx.get("stress_scenarios") is None or ctx["stress_scenarios"].empty:
        return
    ws = wb.create_sheet("Stress Scenarios")
    _set_title(ws, 1, "STRESS-TEST RESULTS", span=10)
    df = ctx["stress_scenarios"].reset_index()
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=3, column=j, value=col)
        c.font, c.fill = FONT_HEADER, FILL_HEADER
    for i, (_, row) in enumerate(df.iterrows(), start=4):
        for j, col in enumerate(df.columns, start=1):
            ws.cell(row=i, column=j, value=row[col])
        # Color band
        band = row.get("band", "")
        ws.cell(row=i, column=df.columns.get_loc("band") + 1).fill = _fill_for_band(band)
    _autosize_columns(ws, max_width=30)


def _sheet_recommendations(wb: Workbook, ctx: dict):
    if not ctx.get("recommendations"):
        return
    ws = wb.create_sheet("Recommendations")
    _set_title(ws, 1, "RULE-BASED RECOMMENDATIONS", span=5)
    headers = ["Priority", "Component", "Action", "Detail", "Expected Impact"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font, c.fill = FONT_HEADER, FILL_HEADER
    for i, r in enumerate(ctx["recommendations"]["actions"], start=4):
        ws.cell(row=i, column=1, value=r.priority)
        ws.cell(row=i, column=2, value=r.component)
        ws.cell(row=i, column=3, value=r.action).font = FONT_LABEL
        ws.cell(row=i, column=4, value=r.detail).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=5, value=r.expected_impact).alignment = Alignment(wrap_text=True)
    _autosize_columns(ws, max_width=70)


def _sheet_macro(wb: Workbook, ctx: dict):
    if not ctx.get("macro"):
        return
    ws = wb.create_sheet("Macro Snapshot")
    _set_title(ws, 1, "CURRENT MACRO INDICATORS", span=3)
    headers = ["Indicator", "Value", "Notes"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font, c.fill = FONT_HEADER, FILL_HEADER
    macro = ctx["macro"]
    rows = [
        ("Inflation (CPI YoY)",   f"{macro.get('inflation_rate', 0):.2%}",   "Headline CPI year-over-year"),
        ("Federal Funds Rate",    f"{macro.get('fed_funds_rate', 0):.2%}",   "Effective FFR"),
        ("Unemployment Rate",     f"{macro.get('unemployment_rate', 0):.2%}", "U3"),
        ("30y Mortgage Rate",     f"{macro.get('mortgage_rate', 0):.2%}",    "Freddie Mac PMMS"),
        ("10y Treasury Yield",    f"{macro.get('treasury_10y', 0):.2%}",     ""),
        ("2y Treasury Yield",     f"{macro.get('treasury_2y', 0):.2%}",      ""),
        ("Yield Curve Spread",    f"{macro.get('yield_curve_spread', 0):.2%}", "10y - 2y"),
        ("VIX",                   f"{macro.get('vix', 0):.1f}",              "Implied volatility (S&P 500)"),
    ]
    for i, (name, val, note) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=name).font = FONT_LABEL
        ws.cell(row=i, column=2, value=val).font = FONT_BODY
        ws.cell(row=i, column=3, value=note).font = Font(name="Calibri", size=10, italic=True)
    _autosize_columns(ws)


# --------------------------------------------------------------------------- #
# Public entry point
# --------------------------------------------------------------------------- #
def generate_report(
    output_path: str | Path,
    fragility_result,
    investment_plan,
    macro: dict,
    portfolio_choice: str,
    initial_capital: float,
    monthly_contribution: float,
    horizon_years: int,
    portfolio_comparison: Optional[pd.DataFrame] = None,
    trade_signals: Optional[List] = None,
    stress_scenarios: Optional[pd.DataFrame] = None,
    recommendations: Optional[Dict] = None,
) -> Path:
    """Generate the full Excel report. Returns the saved file path."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    ctx = {
        "fragility_result": fragility_result,
        "hffi": fragility_result.score,
        "band": fragility_result.band,
        "distress_prob": fragility_result.distress_probability,
        "investment_plan": investment_plan,
        "macro": macro,
        "portfolio_choice": portfolio_choice,
        "initial_capital": initial_capital,
        "monthly_contribution": monthly_contribution,
        "horizon_years": horizon_years,
        "portfolio_comparison": portfolio_comparison,
        "trade_signals": trade_signals,
        "stress_scenarios": stress_scenarios,
        "recommendations": recommendations,
    }

    wb = Workbook()
    wb.remove(wb.active)
    _sheet_summary(wb, ctx)
    _sheet_components(wb, ctx)
    _sheet_investment_plan(wb, ctx)
    _sheet_portfolio_comparison(wb, ctx)
    _sheet_trade_signals(wb, ctx)
    _sheet_stress(wb, ctx)
    _sheet_recommendations(wb, ctx)
    _sheet_macro(wb, ctx)

    wb.save(output_path)
    logger.info("Report saved to %s", output_path)
    return output_path
